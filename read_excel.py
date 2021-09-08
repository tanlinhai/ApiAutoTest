#coding=utf8
import openpyxl
base_path = '/Users/tanlinhai/PycharmProjects/pythonProject2/CaseAdmin/test_case.xlsx'
print(base_path)

'''
xls = openpyxl.load_workbook(base_path + '/CaseAdmin/test_case.xlsx')
sheets = xls.sheetnames
test_sheet = xls[sheets[0]]
print(test_sheet.cell(1, 1).value)
print(test_sheet.max_row)
'''

class HandExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        open_excel = openpyxl.load_workbook(base_path)
        return open_excel

    def get_sheet_data(self, index=None):
        '''
        获取sheet内容
        '''
        sheet_data = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_data[index]]
        return data

    def get_cell_value(self, row, cols):
        '''
        获取某个单元格内容
        '''
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    def get_rows(self):
        '''
        获取行数
        '''
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self, row):
        '''
        获取某行内容
        '''
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        '''
        数据回写
        '''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save('/Users/tanlinhai/PycharmProjects/pythonProject2/CaseAdmin/test_case.xlsx')

excel_data = HandExcel()
if __name__ == '__main__':
    handle_Excel = HandExcel()
    handle_Excel.excel_write_data(2, 11, '通过')